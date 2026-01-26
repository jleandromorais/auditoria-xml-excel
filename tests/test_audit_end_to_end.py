import os
from pathlib import Path
import pandas as pd
from openpyxl import load_workbook

from auditoria.audit import auditar_pasta_pai


def _write_minimal_excel(path: Path, notas):
    """
    Cria um Excel mínimo compatível com excel_loader.py:
    - sheet name contém "25" e "OUT"
    - linha de cabeçalho contém "NOTA" e "S/TRIBUTOS"
    """
    # Cabeçalho (linha 0)
    header = ["NOTA", "S/TRIBUTOS", "VOL", "ICMS", "PIS", "COFINS"]
    rows = [header]
    for nf, liq, vol, icms, pis, cof in notas:
        rows.append([nf, liq, vol, icms, pis, cof])

    df = pd.DataFrame(rows)

    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, header=False, sheet_name="25_OUT")


def _write_nfe_xml(path: Path, nNF: str, vNF="100.00", vICMS="10.00", vPIS="1.00", vCOFINS="2.00", vol="3.000"):
    xml = f"""<?xml version="1.0" encoding="utf-8"?>
<nfeProc xmlns="http://www.portalfiscal.inf.br/nfe">
  <NFe>
    <infNFe>
      <ide>
        <nNF>{nNF}</nNF>
      </ide>
      <total>
        <ICMSTot>
          <vNF>{vNF}</vNF>
          <vICMS>{vICMS}</vICMS>
          <vPIS>{vPIS}</vPIS>
          <vCOFINS>{vCOFINS}</vCOFINS>
        </ICMSTot>
      </total>
      <transp>
        <vol>
          <qVol>{vol}</qVol>
        </vol>
      </transp>
    </infNFe>
  </NFe>
</nfeProc>
"""
    path.write_text(xml, encoding="utf-8")


def test_auditoria_end_to_end(tmp_path: Path):
    # ===== Arrange: estrutura de pastas =====
    pasta_pai = tmp_path / "auditoria_teste"
    emp_a = pasta_pai / "EMPRESA_A"
    emp_b = pasta_pai / "EMPRESA_B"
    emp_a.mkdir(parents=True)
    emp_b.mkdir(parents=True)

    # XML 100 (tem no Excel) -> deve dar OK
    _write_nfe_xml(emp_a / "nf_100.xml", "100", vNF="100.00", vICMS="10.00", vPIS="1.00", vCOFINS="2.00", vol="3.000")

    # XML 200 (não tem no Excel) -> SEM EXCEL
    _write_nfe_xml(emp_b / "nf_200.xml", "200", vNF="50.00", vICMS="5.00", vPIS="0.50", vCOFINS="1.00", vol="1.000")

    # Excel contém 100 (bate) e 300 (não existe XML) -> SEM XML
    excel_path = tmp_path / "base.xlsx"
    notas_excel = [
        ("100", 87.00, 3.000, 10.00, 1.00, 2.00),   # liq esperado = 100-10-1-2 = 87
        ("300", 10.00, 0.500, 0.00, 0.00, 0.00),    # SEM XML
    ]
    _write_minimal_excel(excel_path, notas_excel)

    saida = tmp_path / "saida.xlsx"

    # ===== Act =====
    out_path = auditar_pasta_pai(
        pasta_pai=pasta_pai,
        empresas=[emp_a, emp_b],
        excel_path=str(excel_path),
        saida=str(saida),
    )

    # ===== Assert: arquivo gerado =====
    assert Path(out_path).exists()

    wb = load_workbook(out_path)
    sheet_names = set(wb.sheetnames)

    # Deve existir a aba principal e as duas abas pedidas
    assert "Resultado" in sheet_names
    assert "Excel_sem_XML" in sheet_names
    assert "XML_sem_Excel" in sheet_names

    # Lê Resultado e verifica status
    df_res = pd.read_excel(out_path, sheet_name="Resultado")

    # 100 deve existir e ser OK
    row_100 = df_res[df_res["Nota"].astype(str) == "100"]
    assert not row_100.empty
    assert "OK" in str(row_100.iloc[0]["Status"])

    # 200 deve estar como SEM EXCEL
    row_200 = df_res[df_res["Nota"].astype(str) == "200"]
    assert not row_200.empty
    assert "SEM EXCEL" in str(row_200.iloc[0]["Status"])

    # 300 deve estar como SEM XML
    row_300 = df_res[df_res["Nota"].astype(str) == "300"]
    assert not row_300.empty
    assert "SEM XML" in str(row_300.iloc[0]["Status"])
