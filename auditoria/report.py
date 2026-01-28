import os
import pandas as pd
from datetime import datetime
from typing import List, Dict, Optional
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

from .gemini_writer import gerar_texto_pdf_com_gemini

# Tenta importar reportlab para PDF. Se não tiver, avisa no console.
try:
    from reportlab.lib.pagesizes import letter
    from reportlab.pdfgen import canvas
    HAS_REPORTLAB = True
except ImportError:
    HAS_REPORTLAB = False
    print("Aviso: Biblioteca 'reportlab' não encontrada. PDF não será gerado.")

def gerar_relatorio(lista: List[Dict], saida: Optional[str] = None) -> str:
    """
    Gera o relatório principal (Excel Bonito) e o PDF explicativo.
    """
    if not lista:
        return ""

    df = pd.DataFrame(lista)
    
    # Colunas padrão para garantir a ordem no Excel:
    # Arquivo, Tipo, Nota, Empresa, Volume, valores/impostos, diferenças, e só no final Status/Obs.
    cols_order = [
        "Arquivo",
        "Tipo",
        "Nota",
        "Empresa",
        "Mes",
        "Vol Excel",
        "Vol XML",
        "Diff Vol",
        "Liq Excel",
        "Liq XML (Calc)",
        "ICMS Excel",
        "ICMS XML",
        "PIS Excel",
        "PIS",
        "COFINS Excel",
        "COFINS",
        "Diff R$",
        "Status",
        "Obs",
    ]
    # Garante que todas colunas existem
    for c in cols_order:
        if c not in df.columns:
            df[c] = "-"
    
    df = df[cols_order]

    # Define nome do arquivo se não passar
    if not saida:
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        pasta_relatorios = os.path.join(os.getcwd(), "relatorios")
        os.makedirs(pasta_relatorios, exist_ok=True)
        saida = os.path.join(pasta_relatorios, f"Auditoria_Resultado_{ts}.xlsx")

    # 1. GERA EXCEL ESTILIZADO
    wb = Workbook()
    ws = wb.active
    ws.title = "Resultado Auditoria"

    # Adiciona dados
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)

    # Aplica Estilos "Bonitos"
    _estilizar_planilha(ws)
    
    wb.save(saida)
    print(f"Excel gerado com sucesso: {saida}")

    # 2. GERA PDF EXPLICATIVO (Se possível)
    if HAS_REPORTLAB:
        caminho_pdf = saida.replace(".xlsx", ".pdf")
        _gerar_pdf_resumo(caminho_pdf, df)

    return saida

def gerar_relatorio_avisos(df_duplicadas: pd.DataFrame, lista_sem_xml: List[Dict], caminho_resultado: str) -> str:
    """
    Gera um relatório de AVISOS (separado) bem formatado.
    """
    pasta = os.path.dirname(caminho_resultado)
    nome_base = os.path.basename(caminho_resultado).replace("Resultado", "AVISOS")
    caminho_avisos = os.path.join(pasta, "AVISOS", nome_base)
    
    # Cria pasta específica para avisos para ficar organizado
    os.makedirs(os.path.dirname(caminho_avisos), exist_ok=True)

    wb = Workbook()
    
    # --- ABA 1: DUPLICADAS ---
    if not df_duplicadas.empty:
        ws1 = wb.active
        ws1.title = "Duplicadas no Excel"
        
        # Seleciona colunas úteis
        cols_dup = ["Mes", "NF_Clean", "Vol_Excel", "Liq_Excel", "ICMS_Excel"]
        cols_existentes = [c for c in cols_dup if c in df_duplicadas.columns]
        df_export = df_duplicadas[cols_existentes].sort_values(by="NF_Clean")
        
        for r in dataframe_to_rows(df_export, index=False, header=True):
            ws1.append(r)
        
        _estilizar_planilha(ws1, cor_padrao="FFFFE0") # Amarelo claro para avisos

    # --- ABA 2: SEM XML ---
    if lista_sem_xml:
        if "Duplicadas no Excel" in wb.sheetnames:
            ws2 = wb.create_sheet("Faltam XMLs")
        else:
            ws2 = wb.active
            ws2.title = "Faltam XMLs"

        df_sem = pd.DataFrame(lista_sem_xml)
        cols_sem = ["Nota", "Mes", "Liq Excel", "Status", "Obs"]
        cols_sem = [c for c in cols_sem if c in df_sem.columns]
        
        for r in dataframe_to_rows(df_sem[cols_sem], index=False, header=True):
            ws2.append(r)

        _estilizar_planilha(ws2, cor_padrao="FFC7CE") # Vermelho claro para erros

    wb.save(caminho_avisos)
    return caminho_avisos

def _estilizar_planilha(ws, cor_padrao=None):
    """Aplica formatação profissional: Cabeçalho Azul, Zebra ou Condicional."""
    # Cores
    header_fill = PatternFill(start_color="203764", end_color="203764", fill_type="solid") # Azul Escuro
    header_font = Font(bold=True, color="FFFFFF")
    
    fill_verde = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    font_verde = Font(color="006100")
    
    fill_vermelho = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    font_vermelho = Font(color="9C0006")

    fill_amarelo = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    font_amarelo = Font(color="9C6500")

    # Formata Cabeçalho
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Tenta achar coluna de Status
    col_status = None
    for cell in ws[1]:
        if str(cell.value).lower() == "status":
            col_status = cell.column
            break

    # Formata Linhas
    for row in ws.iter_rows(min_row=2):
        status_val = ""
        if col_status:
            status_val = str(row[col_status-1].value).upper()

        fill_atual = None
        font_atual = None

        if "OK" in status_val:
            fill_atual = fill_verde
            font_atual = font_verde
        elif "ERRO" in status_val or "SEM" in status_val:
            fill_atual = fill_vermelho
            font_atual = font_vermelho
        elif "PENDENTE" in status_val:
            fill_atual = fill_amarelo
            font_atual = font_amarelo
        elif cor_padrao:
             fill_atual = PatternFill(start_color=cor_padrao, end_color=cor_padrao, fill_type="solid")

        if fill_atual:
            for cell in row:
                cell.fill = fill_atual
                if font_atual: cell.font = font_atual
        
        # Formatação de Números
        for cell in row:
            v = cell.value
            if isinstance(v, (int, float)):
                cell.number_format = '#,##0.00'

    # Ajuste Automático de Largura
    for col in ws.columns:
        largura = 15
        try:
            val = str(col[0].value)
            if len(val) > largura: largura = len(val) + 2
        except: pass
        ws.column_dimensions[col[0].column_letter].width = largura

def _gerar_pdf_resumo(caminho, df):
    """Gera PDF com a explicação solicitada."""
    c = canvas.Canvas(caminho, pagesize=letter)
    w, h = letter
    
    # Cabeçalho
    c.setFont("Helvetica-Bold", 16)
    c.drawString(50, h - 50, "Relatório de Auditoria - Resumo Executivo")
    c.setFont("Helvetica", 10)
    c.drawString(50, h - 70, f"Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M')}")
    
    # Resumo Numérico
    total = len(df)
    oks = len(df[df["Status"].astype(str).str.contains("OK")])
    erros = len(df[df["Status"].astype(str).str.contains("ERRO")])
    sem_xml = len(df[df["Status"].astype(str).str.contains("SEM XML")])

    y = h - 110
    c.setFont("Helvetica-Bold", 12)
    c.drawString(50, y, "Estatísticas:")
    y -= 20
    c.setFont("Helvetica", 10)
    c.drawString(60, y, f"Total Analisado: {total}")
    c.drawString(60, y-15, f"Sucesso (OK): {oks}")
    c.drawString(60, y-30, f"Divergências de Valor/Vol: {erros}")
    c.drawString(60, y-45, f"Sem XML correspondente: {sem_xml}")

    # TEXTO DO PDF (IA opcional com fallback)
    y -= 80
    c.setFont("Helvetica-Bold", 12)
    c.drawString(50, y, "Resumo Executivo e Nota Técnica")

    texto_ia = gerar_texto_pdf_com_gemini(df)
    if texto_ia:
        texto = texto_ia
    else:
        texto = [
            "Foi observado que a planilha Excel ('Conta Gráfica') possui mais registros que a quantidade",
            "de arquivos XML processados. Isso ocorre pelos seguintes motivos identificados:",
            "",
            "1. Notas de Débito (ND): O Excel contém lançamentos financeiros de ajustes (NDs) que",
            "   não são Notas Fiscais eletrônicas e não possuem arquivo XML padrão.",
            "2. Penalidades/Multas: Cobranças como 'Falha de Programação' são lançadas no Excel",
            "   mas documentadas via boletos/PDFs simples, sem XML processável pelo sistema.",
            "3. Documentos em PDF: O sistema auditou apenas arquivos .XML. Documentos salvos",
            "   apenas como PDF não são contabilizados na coluna 'Vol XML'.",
            "",
            "Recomendação: Verifique a pasta 'AVISOS' gerada junto a este relatório para ver",
            "a lista exata dos itens que constam no Excel mas não tiveram XML encontrado."
        ]
    
    y -= 20
    c.setFont("Helvetica", 10)
    for linha in texto:
        # Quebra de página simples
        if y < 60:
            c.showPage()
            c.setFont("Helvetica", 10)
            y = h - 60
        c.drawString(50, y, linha)
        y -= 15

    c.save()
    print(f"PDF Explicativo salvo: {caminho}")