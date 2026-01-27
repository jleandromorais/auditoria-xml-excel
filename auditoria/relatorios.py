import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from datetime import datetime
import os

class GeradorRelatorios:
    def __init__(self, dados_auditoria, pasta_saida="auditoria/relatorios"):
        """
        dados_auditoria: Lista de dicionários contendo os resultados da comparação.
        Exemplo: [{'arquivo': 'NF123.xml', 'status': 'OK', 'detalhe': 'Valores batem'}, ...]
        """
        self.dados = dados_auditoria
        self.pasta_saida = pasta_saida
        if not os.path.exists(pasta_saida):
            os.makedirs(pasta_saida)

    def gerar_excel_estilizado(self):
        df = pd.DataFrame(self.dados)
        nome_arquivo = f"{self.pasta_saida}/Relatorio_Auditoria_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Auditoria"

        # Estilos
        fill_verde = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid") # Fundo Verde Claro
        font_verde = Font(color="006100") # Texto Verde Escuro
        
        fill_vermelho = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid") # Fundo Vermelho Claro
        font_vermelho = Font(color="9C0006") # Texto Vermelho Escuro
        
        fill_amarelo = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid") # Fundo Amarelo
        font_amarelo = Font(color="9C6500") # Texto Amarelo Escuro

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid") # Azul bonito

        # Adiciona cabeçalho
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)

        # Formatação do Cabeçalho
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        # Formatação Condicional das Linhas
        # Assume que a coluna 'status' ou 'divergencia' é a coluna B (índice 2 no openpyxl)
        coluna_status = 2 
        
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            status_cell = row[coluna_status - 1] # Ajuste o índice conforme sua coluna de status
            valor_status = str(status_cell.value).lower()

            estilo_fill = None
            estilo_font = None

            if "ok" in valor_status or "sucesso" in valor_status:
                estilo_fill = fill_verde
                estilo_font = font_verde
            elif "erro" in valor_status or "faltante" in valor_status or "divergencia" in valor_status:
                estilo_fill = fill_vermelho
                estilo_font = font_vermelho
            elif "aviso" in valor_status:
                estilo_fill = fill_amarelo
                estilo_font = font_amarelo

            if estilo_fill:
                for cell in row:
                    cell.fill = estilo_fill
                    cell.font = estilo_font

        # Ajustar largura das colunas
        for column in ws.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column[0].column_letter].width = adjusted_width

        wb.save(nome_arquivo)
        print(f"Excel estilizado salvo em: {nome_arquivo}")

    def gerar_pdf_explicativo(self):
        nome_arquivo = f"{self.pasta_saida}/Resumo_Auditoria_{datetime.now().strftime('%Y%m%d_%H%M')}.pdf"
        c = canvas.Canvas(nome_arquivo, pagesize=letter)
        width, height = letter

        # Cabeçalho
        c.setFont("Helvetica-Bold", 16)
        c.drawString(50, height - 50, "Relatório de Auditoria XML vs Excel")
        
        c.setFont("Helvetica", 10)
        c.drawString(50, height - 70, f"Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M')}")

        # Resumo
        total_analisado = len(self.dados)
        c.drawString(50, height - 100, f"Total de itens analisados: {total_analisado}")

        # A Explicação Técnica (Sua solicitação)
        c.line(50, height - 120, width - 50, height - 120)
        
        c.setFont("Helvetica-Bold", 12)
        c.drawString(50, height - 140, "Nota Técnica: Discrepância de Quantidade (Excel > XML)")
        
        texto_explicativo = [
            "Durante a auditoria, é comum observar que a planilha Excel contém mais registros",
            "do que a quantidade de arquivos XML processados. Isso ocorre pelos motivos:",
            "",
            "1. Arquivos Faltantes: O registro existe no Excel (financeiro), mas o arquivo",
            "   digital (XML) não foi encontrado na pasta de origem.",
            "2. Notas de Débito/Crédito: Ajustes financeiros lançados no Excel que não",
            "   possuem um XML de Nota Fiscal correspondente.",
            "3. Documentos PDF: Faturas apenas em PDF não são lidas pelo processador de XML.",
            "",
            "Recomendação: Verifique a pasta 'Avisos' no Excel gerado para identificar",
            "quais notas específicas estão faltando na pasta digital."
        ]

        y = height - 160
        c.setFont("Helvetica", 10)
        for linha in texto_explicativo:
            c.drawString(50, y, linha)
            y -= 15

        c.save()
        print(f"PDF explicativo salvo em: {nome_arquivo}")

# Exemplo de como testar/usar
if __name__ == "__main__":
    # Dados de teste para você ver como fica
    dados_teste = [
        {"Arquivo": "NF001.xml", "Status": "OK", "Valor Excel": 100, "Valor XML": 100},
        {"Arquivo": "NF002.xml", "Status": "Erro - Valor Divergente", "Valor Excel": 150, "Valor XML": 140},
        {"Arquivo": "NF003.xml", "Status": "Aviso - XML não encontrado", "Valor Excel": 200, "Valor XML": 0},
    ]
    
    gerador = GeradorRelatorios(dados_teste)
    gerador.gerar_excel_estilizado()
    gerador.gerar_pdf_explicativo()